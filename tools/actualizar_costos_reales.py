import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
import unicodedata
import re
from difflib import SequenceMatcher

def normalize_text(text):
    if not text or pd.isna(text):
        return ''
    text = str(text).upper()
    text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
    text = re.sub(r'[^A-Z0-9]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def extract_numbers(text):
    return set(re.findall(r'\b\d+(?:[\.,]\d+)?\b', text))

# 1. Cargar CATALOGO_COMPLETO
path_bd = 'EXCELES PRINCIPALES/BASE DE DATOS CONTEO FISICO AGOSTO 2026 (2).xlsx'
df_cat = pd.read_excel(path_bd, sheet_name='CATALOGO_COMPLETO')

cat_items = []
codigos_existentes = set()

for _, row in df_cat.iterrows():
    cod = str(row['CODIGO']).strip() if pd.notna(row['CODIGO']) else ''
    nom = str(row['INSUMO']).strip() if pd.notna(row['INSUMO']) else ''
    if cod and nom:
        codigos_existentes.add(cod)
        norm = normalize_text(nom)
        cat_items.append({
            'codigo': cod,
            'nombre': nom,
            'norm': norm,
            'tokens': set(norm.split()),
            'numbers': extract_numbers(norm)
        })

print(f'Total catálogo preparado: {len(cat_items)}')

def find_best_match(nombre_buscado):
    norm_q = normalize_text(nombre_buscado)
    if not norm_q or len(norm_q) < 3:
        return None, 0.0, 'CREAR NUEVO'
    
    tokens_q = set(norm_q.split())
    numbers_q = extract_numbers(norm_q)
    
    best_item = None
    best_score = 0.0
    
    for item in cat_items:
        # 1. Coincidencia exacta
        if norm_q == item['norm']:
            return item, 1.0, 'EXACTO'
            
        # 2. Verificar discrepancia de números (medidas, onzas, tamaños)
        if numbers_q and item['numbers']:
            # Si ambos tienen números y no comparten ninguno de los números clave
            if not (numbers_q & item['numbers']):
                continue  # No emparejar si los tamaños/números son distintos
                
        inter = len(tokens_q & item['tokens'])
        union = len(tokens_q | item['tokens'])
        jaccard = inter / union if union > 0 else 0
        
        ratio = SequenceMatcher(None, norm_q, item['norm']).ratio()
        containment = inter / len(tokens_q) if tokens_q else 0
        
        score = (containment * 0.45) + (ratio * 0.35) + (jaccard * 0.20)
        
        tokens_q_list = norm_q.split()
        tokens_item_list = item['norm'].split()
        if tokens_q_list and tokens_item_list and tokens_q_list[0] == tokens_item_list[0]:
            score += 0.10
            
        if score > best_score:
            best_score = score
            best_item = item
            
    if best_score >= 0.70 and best_item:
        return best_item, best_score, 'EMPAREJADO'
    else:
        return None, best_score, 'CREAR NUEVO'

# 2. Cargar y Actualizar COSTOS REALES.xlsx con openpyxl
path_costos = 'EXCELES PRINCIPALES/COSTOS REALES.xlsx'
wb = openpyxl.load_workbook(path_costos)
ws = wb['CAT2026']

# Encabezados en fila 6
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

ws.cell(row=6, column=8, value='CODIGO_ASIGNADO').font = header_font
ws.cell(row=6, column=8).fill = header_fill
ws.cell(row=6, column=8).alignment = Alignment(horizontal='center', vertical='center')

ws.cell(row=6, column=9, value='INSUMO_CATALOGO_SUGERIDO').font = header_font
ws.cell(row=6, column=9).fill = header_fill
ws.cell(row=6, column=9).alignment = Alignment(horizontal='center', vertical='center')

ws.cell(row=6, column=10, value='ESTADO_ASOCIACION').font = header_font
ws.cell(row=6, column=10).fill = header_fill
ws.cell(row=6, column=10).alignment = Alignment(horizontal='center', vertical='center')

fill_exacto = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid') # Verde claro
fill_emparejado = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid') # Azul claro
fill_nuevo = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid') # Rojo claro

total_filas = 0
con_codigo_previo = 0
emparejados_auto = 0
marcados_crear_nuevo = 0

for row in range(7, ws.max_row + 1):
    cod_val = ws.cell(row=row, column=1).value
    nombre_val = ws.cell(row=row, column=2).value
    
    if not nombre_val or str(nombre_val).strip() == '':
        continue
        
    total_filas += 1
    nombre_str = str(nombre_val).strip()
    
    # Caso 1: Ya tiene código en el archivo original
    if cod_val and str(cod_val).strip() != '' and str(cod_val).strip().lower() != 'nan':
        cod_str = str(cod_val).strip()
        ws.cell(row=row, column=8, value=cod_str).alignment = Alignment(horizontal='center')
        
        # Buscar nombre en catálogo si existe
        nom_cat = next((item['nombre'] for item in cat_items if item['codigo'] == cod_str), nombre_str)
        ws.cell(row=row, column=9, value=nom_cat)
        ws.cell(row=row, column=10, value='CODIGO PREVIO').fill = fill_exacto
        ws.cell(row=row, column=10).font = Font(color='065F46', bold=True, size=10)
        con_codigo_previo += 1
        
    # Caso 2: No tiene código -> Ejecutar emparejamiento inteligente
    else:
        best_item, score, status = find_best_match(nombre_str)
        if status in ('EMPAREJADO', 'EXACTO') and best_item:
            ws.cell(row=row, column=8, value=best_item['codigo']).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=8).font = Font(bold=True, color='1E40AF')
            ws.cell(row=row, column=9, value=best_item['nombre'])
            ws.cell(row=row, column=10, value=f'EMPAREJADO ({int(score*100)}%)').fill = fill_emparejado
            ws.cell(row=row, column=10).font = Font(color='1E40AF', bold=True, size=10)
            emparejados_auto += 1
        else:
            ws.cell(row=row, column=8, value='CREAR NUEVO').alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=8).font = Font(bold=True, color='991B1B')
            ws.cell(row=row, column=9, value='-')
            ws.cell(row=row, column=10, value='CREAR NUEVO').fill = fill_nuevo
            ws.cell(row=row, column=10).font = Font(color='991B1B', bold=True, size=10)
            marcados_crear_nuevo += 1

# Ajustar anchos de columnas H, I, J
ws.column_dimensions['H'].width = 18
ws.column_dimensions['I'].width = 45
ws.column_dimensions['J'].width = 22

# Guardar archivo actualizado
wb.save('EXCELES PRINCIPALES/COSTOS REALES.xlsx')
print('=== PROCESO COMPLETADO EXITOSAMENTE ===')
print(f'Total filas procesadas: {total_filas}')
print(f'Filas con código previo mantenido: {con_codigo_previo}')
print(f'Filas emparejadas automáticamente por nombre: {emparejados_auto}')
print(f'Filas marcadas como CREAR NUEVO: {marcados_crear_nuevo}')
print('Archivo COSTOS REALES.xlsx guardado y actualizado.')
