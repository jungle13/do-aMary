import flet as ft
from config import Config
from core.supabase_client import SupabaseClient
import datetime
from calendar import monthrange
import threading
from core.fecha_utils import parsear_a_fecha_local, formatear_fecha_hora_local, get_ahora_local, get_hoy_local_str

class InformesView(ft.Container):
    def __init__(self):
        super().__init__()
        self.expand = True
        self.db = SupabaseClient()
        self.save_pdf_picker = ft.FilePicker(on_result=self._save_pdf_result)
        self.save_excel_picker = ft.FilePicker(on_result=self._save_excel_result)
        
        # --- PANEL IZQUIERDO: CONSTRUCTOR DE INFORMES ---
        self.drop_tipo_informe = ft.Dropdown(
            label="Tipo de Informe",
            options=[
                ft.dropdown.Option("Valorización de Inventario"),
                ft.dropdown.Option("Informe de Compras"),
                ft.dropdown.Option("Informe de Ventas"),
                ft.dropdown.Option("Historial de Ajustes"),
                ft.dropdown.Option("Informe de Impuestos"),
                ft.dropdown.Option("Informe de Recaudos"),
                ft.dropdown.Option("Resumen de KPIs")
            ],
            value="Valorización de Inventario",
            on_change=self._on_tipo_informe_change,
            dense=True, border_radius=8,
            height=38, text_size=12,
            content_padding=ft.padding.symmetric(horizontal=10, vertical=8)
        )

        self.drop_agrupacion_ventas = ft.Dropdown(
            label="Agrupar Ventas por",
            options=[
                ft.dropdown.Option("Tipo de Documento (POS / Remisión)"),
                ft.dropdown.Option("Categoría"),
                ft.dropdown.Option("Documento y Categoría")
            ],
            value="Tipo de Documento (POS / Remisión)",
            dense=True, border_radius=8,
            height=38, text_size=12,
            content_padding=ft.padding.symmetric(horizontal=10, vertical=8),
            visible=False
        )
        
        self.drop_filtro_fecha = ft.Dropdown(
            label="Periodo para Informe / PDF",
            options=[],
            value="HISTORICO",
            dense=True, border_radius=8,
            height=38, text_size=12,
            content_padding=ft.padding.symmetric(horizontal=10, vertical=8)
        )
        self._cargar_opciones_periodo()
        
        self.opcion_detalle = ft.RadioGroup(
            content=ft.Row([
                ft.Radio(value="Completo", label="Completo"),
                ft.Radio(value="Resumido", label="Resumido")
            ]),
            value="Completo"
        )
        
        self.btn_generar = ft.ElevatedButton(
            "Generar Previsualización", 
            icon=ft.icons.PLAY_ARROW, 
            bgcolor=Config.COLOR_PRIMARY, 
            color="white",
            on_click=self.generar_informe,
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )
        
        self.btn_pdf = ft.OutlinedButton(
            "Exportar a PDF (Periodo Seleccionado)", 
            icon=ft.icons.PICTURE_AS_PDF, 
            icon_color="red", 
            on_click=self.exportar_pdf,
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )
        
        self.btn_excel = ft.OutlinedButton(
            "Exportar a Excel (Consolidado General)", 
            icon=ft.icons.TABLE_VIEW, 
            icon_color="green", 
            on_click=self.exportar_excel,
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )
        
        # Tarjeta informativa corta para la exportación a Excel
        info_excel_box = ft.Container(
            content=ft.Row([
                ft.Icon(ft.icons.INFO_OUTLINED, size=15, color="blue700"),
                ft.Text(
                    "Excel genera el consolidado general completo (Inventario, Compras, Ventas y Ajustes) acumulado a la fecha, sin aplicar los filtros seleccionados.",
                    size=10,
                    color="blue900",
                    expand=True
                )
            ], alignment=ft.MainAxisAlignment.START, vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=6),
            padding=8,
            bgcolor="#e3f2fd",
            border_radius=6,
            border=ft.border.all(1, "#bbdefb")
        )
        
        panel_controles = ft.Container(
            content=ft.Column([
                ft.Text("Parámetros del Informe", weight="bold", color=Config.COLOR_PRIMARY),
                ft.Divider(height=1, color="#eeeeee"),
                self.drop_tipo_informe,
                self.drop_agrupacion_ventas,
                self.drop_filtro_fecha,
                ft.Text("Nivel de Detalle", size=12, color="grey"),
                self.opcion_detalle,
                ft.Container(height=10),
                self.btn_generar,
                ft.Divider(height=15, color="transparent"),
                ft.Text("Exportación", weight="bold", color=Config.COLOR_PRIMARY),
                ft.Divider(height=1, color="#eeeeee"),
                self.btn_pdf,
                ft.Divider(height=8, color="#f0f0f0"), # Separador suave entre PDF y Excel
                info_excel_box,
                self.btn_excel
            ], spacing=12),
            bgcolor="white", padding=20, border_radius=8, width=300,
            border=ft.border.all(1, "#e0e0e0"),
            shadow=ft.BoxShadow(spread_radius=1, blur_radius=5, color=ft.colors.with_opacity(0.05, "black"))
        )
        
        # --- PANEL DERECHO: LIENZO DEL DOCUMENTO (A4) ---
        self.doc_header_empresa = ft.Text("TIENDA Y ABARROTES LOS DESECHABLES DE DOÑA MARY SAS", weight="bold", size=16, text_align=ft.TextAlign.CENTER)
        self.doc_header_titulo = ft.Text("INFORME DE VALORIZACIÓN DE INVENTARIO", weight="bold", size=14, text_align=ft.TextAlign.CENTER)
        self.doc_header_periodo = ft.Text("Periodo: Histórico Completo", size=11, color="grey", text_align=ft.TextAlign.CENTER)
        self.doc_header_fecha = ft.Text(f"Fecha de Generación: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", size=11, color="grey", text_align=ft.TextAlign.CENTER)
        
        self.doc_cuerpo = ft.Column(spacing=5)
        
        self.lienzo_documento = ft.Container(
            content=ft.Column([
                ft.Container(
                    content=ft.Column([
                        self.doc_header_empresa,
                        self.doc_header_titulo,
                        self.doc_header_periodo,
                        self.doc_header_fecha
                    ], spacing=2, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                    alignment=ft.alignment.center,
                    padding=ft.padding.only(bottom=20)
                ),
                ft.Divider(height=2, color="black"),
                self.doc_cuerpo
            ]),
            bgcolor="white",
            padding=40,
            border_radius=5,
            shadow=ft.BoxShadow(spread_radius=2, blur_radius=10, color=ft.colors.with_opacity(0.1, "black"))
        )
        
        scroll_lienzo = ft.Column([self.lienzo_documento], scroll=ft.ScrollMode.ALWAYS, expand=True)
        
        # --- ENSAMBLAJE FINAL ---
        self.content = ft.Row([
            panel_controles,
            scroll_lienzo
        ], expand=True, spacing=20, vertical_alignment=ft.CrossAxisAlignment.START)

    def _cargar_opciones_periodo(self):
        try:
            from core.periodo_manager import PeriodoManager
            pm = PeriodoManager()
            periodos = pm.periodos_disponibles or pm.cargar_periodos()
        except Exception:
            periodos = []
        
        meses_nombres = {
            "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
            "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
            "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre"
        }
        
        opciones = [
            ft.dropdown.Option(key="HISTORICO", text="Histórico Completo"),
            ft.dropdown.Option(key="MES_ACTUAL", text="Mes Actual"),
            ft.dropdown.Option(key="HOY", text="Día de Hoy")
        ]
        
        for p in periodos:
            mes_cod = p.get("mes_periodo", "")
            if mes_cod and len(mes_cod) == 7 and "-" in mes_cod:
                partes = mes_cod.split("-")
                nom_mes = meses_nombres.get(partes[1], partes[1])
                anio = partes[0]
                label_p = f"{nom_mes} {anio} ({mes_cod})"
                opciones.append(ft.dropdown.Option(key=mes_cod, text=label_p))
                
        self.drop_filtro_fecha.options = opciones
        if not self.drop_filtro_fecha.value:
            self.drop_filtro_fecha.value = "HISTORICO"

    def _on_tipo_informe_change(self, e):
        self.drop_agrupacion_ventas.visible = (self.drop_tipo_informe.value == "Informe de Ventas")
        if self.page:
            self.page.update()

    def did_mount(self):
        self._cargar_opciones_periodo()
        if self.page:
            if self.save_pdf_picker not in self.page.overlay:
                self.page.overlay.append(self.save_pdf_picker)
            if self.save_excel_picker not in self.page.overlay:
                self.page.overlay.append(self.save_excel_picker)
            self.page.update()

    def generar_informe(self, e):
        self.doc_cuerpo.controls.clear()
        self.doc_cuerpo.controls.append(ft.Container(content=ft.ProgressRing(), alignment=ft.alignment.center, padding=50))
        if self.page: self.page.update()
        threading.Thread(target=self._worker_generar_informe, daemon=True).start()

    def _worker_generar_informe(self):
        tipo_informe = self.drop_tipo_informe.value
        detalle = self.opcion_detalle.value
        periodo_filtro = self.drop_filtro_fecha.value or "HISTORICO"

        # 1. Cálculo del Rango de Fechas
        hoy = datetime.date.today()
        mes_periodo_kpi = None

        if periodo_filtro in ("HOY", "Día de Hoy"):
            fecha_inicio = hoy.strftime("%Y-%m-%d")
            fecha_fin = hoy.strftime("%Y-%m-%d")
            periodo_label = f"Día de Hoy ({fecha_inicio})"
        elif periodo_filtro in ("MES_ACTUAL", "Mes Actual"):
            fecha_inicio = hoy.replace(day=1).strftime("%Y-%m-%d")
            ultimo_dia = monthrange(hoy.year, hoy.month)[1]
            fecha_fin = hoy.replace(day=ultimo_dia).strftime("%Y-%m-%d")
            periodo_label = f"Mes Actual ({fecha_inicio} al {fecha_fin})"
            mes_periodo_kpi = hoy.strftime("%Y-%m")
        elif periodo_filtro in ("HISTORICO", "Histórico Completo"):
            fecha_inicio = "2000-01-01"
            fecha_fin = "2100-12-31"
            periodo_label = "Histórico Completo"
        elif len(periodo_filtro) == 7 and "-" in periodo_filtro:
            year, month = map(int, periodo_filtro.split("-"))
            ultimo_dia = monthrange(year, month)[1]
            fecha_inicio = f"{periodo_filtro}-01"
            fecha_fin = f"{periodo_filtro}-{ultimo_dia:02d}"
            meses_nombres = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}
            nom_m = meses_nombres.get(month, str(month))
            periodo_label = f"{nom_m} {year} ({fecha_inicio} al {fecha_fin})"
            mes_periodo_kpi = periodo_filtro
        else:
            fecha_inicio = "2000-01-01"
            fecha_fin = "2100-12-31"
            periodo_label = str(periodo_filtro)

        # 2. Actualizar Cabecera del Documento
        tit_fmt = tipo_informe.upper() if tipo_informe.upper().startswith("INFORME") else f"INFORME DE {tipo_informe.upper()}"
        self.doc_header_titulo.value = tit_fmt
        self.doc_header_periodo.value = f"Periodo: {periodo_label} | Tipo: {detalle}"
        self.doc_header_fecha.value = f"Fecha de Generación: {datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')}"

        self.doc_cuerpo.controls.clear()

        # 3. Enrutador según el tipo de informe
        if tipo_informe == "Valorización de Inventario":
            self._generar_valorizacion(detalle, fecha_corte=fecha_fin if periodo_filtro not in ("HISTORICO", "Histórico Completo") else None)
        elif tipo_informe == "Informe de Compras":
            self._generar_compras(fecha_inicio, fecha_fin, detalle)
        elif tipo_informe == "Informe de Ventas":
            self._generar_ventas(fecha_inicio, fecha_fin, detalle)
        elif tipo_informe == "Historial de Ajustes":
            self._generar_ajustes(fecha_inicio, fecha_fin, detalle)
        elif tipo_informe == "Informe de Impuestos":
            self._generar_impuestos(fecha_inicio, fecha_fin, detalle)
        elif tipo_informe == "Informe de Recaudos":
            self._generar_recaudos(fecha_inicio, fecha_fin, detalle)
        elif tipo_informe == "Resumen de KPIs":
            self._generar_kpis(fecha_inicio, fecha_fin, mes_periodo=mes_periodo_kpi)

        if self.page:
            self.page.update()

    def _generar_impuestos(self, fecha_inicio, fecha_fin, detalle):
        raw_compras = self.db._db.get_all("registro_compras?select=*,catalogo_insumos(nombre)&estado_registro=eq.VÁLIDO&order=fecha.asc") or []
        raw_ventas = self.db._db.get_all("registro_ventas?select=*,catalogo_insumos(nombre)&estado_registro=neq.ANULADO&order=fecha.asc") or []

        # Filtrar Compras por Rango de Fechas
        compras_filtradas = []
        tot_compras_base = 0.0
        tot_compras_iva = 0.0
        tot_compras_total = 0.0

        for c in raw_compras:
            f = str(c.get("fecha") or "")[:10]
            if fecha_inicio <= f <= fecha_fin:
                tot = float(c.get("costo_total") or 0)
                iva = float(c.get("iva") or c.get("valor_iva") or 0)
                base = tot - iva
                cant = float(c.get("cantidad") or 0)
                cat_i = c.get("catalogo_insumos") or {}
                
                compras_filtradas.append({
                    "fecha": f,
                    "doc": c.get("numero_factura") or c.get("numero_entrada") or "S/D",
                    "insumo": cat_i.get("nombre", "Desconocido"),
                    "cant": cant,
                    "base": base,
                    "iva": iva,
                    "total": tot
                })
                tot_compras_base += base
                tot_compras_iva += iva
                tot_compras_total += tot

        # Filtrar Ventas por Rango de Fechas
        ventas_filtradas = []
        tot_ventas_base = 0.0
        tot_ventas_iva = 0.0
        tot_ventas_total = 0.0

        for v in raw_ventas:
            f = parsear_a_fecha_local(v.get("fecha"))
            if fecha_inicio <= f <= fecha_fin:
                tot = float(v.get("total") or 0)
                iva = float(v.get("iva") or 0)
                base = float(v.get("subtotal") or (tot - iva))
                cant = float(v.get("cantidad") or 0)
                cat_i = v.get("catalogo_insumos") or {}

                ventas_filtradas.append({
                    "fecha": f,
                    "doc": v.get("factura_no") or "S/D",
                    "insumo": v.get("descripcion") or cat_i.get("nombre") or "Desconocido",
                    "cant": cant,
                    "base": base,
                    "iva": iva,
                    "total": tot
                })
                tot_ventas_base += base
                tot_ventas_iva += iva
                tot_ventas_total += tot

        balance_iva = tot_ventas_iva - tot_compras_iva

        # Guardar estructura para PDF/Excel
        self.current_data = {
            "compras": compras_filtradas,
            "ventas": ventas_filtradas,
            "tot_compras_base": tot_compras_base,
            "tot_compras_iva": tot_compras_iva,
            "tot_compras_total": tot_compras_total,
            "tot_ventas_base": tot_ventas_base,
            "tot_ventas_iva": tot_ventas_iva,
            "tot_ventas_total": tot_ventas_total,
            "balance_iva": balance_iva
        }
        self.current_total = balance_iva
        self.current_periodo = self.doc_header_periodo.value

        def _crear_resumen_impuestos_row(label, val_base, val_iva, val_total, is_header=False, color_txt="black"):
            weight = "bold" if is_header else "normal"
            size = 12 if is_header else 11
            return ft.Row([
                ft.Text(label, weight=weight, size=size, expand=True, color=color_txt),
                ft.Text(f"${val_base:,.2f}" if isinstance(val_base, (int, float)) else val_base, weight=weight, size=size, width=110, text_align=ft.TextAlign.RIGHT, color=color_txt),
                ft.Text(f"${val_iva:,.2f}" if isinstance(val_iva, (int, float)) else val_iva, weight=weight, size=size, width=110, text_align=ft.TextAlign.RIGHT, color=color_txt),
                ft.Text(f"${val_total:,.2f}" if isinstance(val_total, (int, float)) else val_total, weight=weight, size=size, width=120, text_align=ft.TextAlign.RIGHT, color=color_txt),
            ])

        if detalle == "Resumido":
            self.doc_cuerpo.controls.append(_crear_resumen_impuestos_row("CONCEPTO / CONSOLIDADO", "VALOR BASE", "IVA ACUMULADO", "VALOR TOTAL", is_header=True))
            self.doc_cuerpo.controls.append(ft.Divider(height=1, color="black"))
            
            self.doc_cuerpo.controls.append(_crear_resumen_impuestos_row("VENTAS (IVA GENERADO)", tot_ventas_base, tot_ventas_iva, tot_ventas_total, color_txt="blue700"))
            self.doc_cuerpo.controls.append(_crear_resumen_impuestos_row("COMPRAS (IVA PAGADO / DESCONTABLE)", tot_compras_base, tot_compras_iva, tot_compras_total, color_txt="teal700"))
            self.doc_cuerpo.controls.append(ft.Divider(height=2, color="black"))

            color_bal = "red" if balance_iva > 0 else "green"
            lbl_bal = "BALANCE NETO DE IVA (POR PAGAR)" if balance_iva > 0 else "BALANCE NETO DE IVA (A FAVOR)"
            self.doc_cuerpo.controls.append(
                ft.Row([
                    ft.Text(f"{lbl_bal}:", weight="bold", size=13, expand=True, text_align=ft.TextAlign.RIGHT),
                    ft.Text(f"${balance_iva:,.2f}", weight="bold", size=14, width=150, text_align=ft.TextAlign.RIGHT, color=color_bal),
                ])
            )
        else:
            # VISTA COMPLETA DETALLADA
            # 1. SECCIÓN COMPRAS
            self.doc_cuerpo.controls.append(ft.Container(content=ft.Text("COMPRAS (IVA PAGADO EN ENTRADAS)", weight="bold", size=13, color="teal700"), padding=ft.padding.only(top=10, bottom=5)))
            self.doc_cuerpo.controls.append(ft.Row([
                ft.Text("FECHA", weight="bold", size=10, width=65),
                ft.Text("DOC.", weight="bold", size=10, width=80),
                ft.Text("INSUMO", weight="bold", size=10, expand=True),
                ft.Text("BASE", weight="bold", size=10, width=80, text_align=ft.TextAlign.RIGHT),
                ft.Text("IVA", weight="bold", size=10, width=70, text_align=ft.TextAlign.RIGHT),
                ft.Text("TOTAL", weight="bold", size=10, width=85, text_align=ft.TextAlign.RIGHT),
            ]))
            self.doc_cuerpo.controls.append(ft.Divider(height=1, color="black"))

            for c in compras_filtradas:
                self.doc_cuerpo.controls.append(ft.Row([
                    ft.Text(c['fecha'], size=10, width=65),
                    ft.Text(c['doc'], size=10, width=80, no_wrap=True),
                    ft.Text(c['insumo'], size=10, expand=True, no_wrap=True),
                    ft.Text(f"${c['base']:,.2f}", size=10, width=80, text_align=ft.TextAlign.RIGHT),
                    ft.Text(f"${c['iva']:,.2f}", size=10, width=70, text_align=ft.TextAlign.RIGHT),
                    ft.Text(f"${c['total']:,.2f}", size=10, width=85, text_align=ft.TextAlign.RIGHT),
                ]))

            self.doc_cuerpo.controls.append(ft.Row([
                ft.Text("TOTAL COMPRAS:", weight="bold", size=11, expand=True, text_align=ft.TextAlign.RIGHT),
                ft.Text(f"${tot_compras_base:,.2f}", weight="bold", size=11, width=80, text_align=ft.TextAlign.RIGHT),
                ft.Text(f"${tot_compras_iva:,.2f}", weight="bold", size=11, width=70, text_align=ft.TextAlign.RIGHT, color="teal700"),
                ft.Text(f"${tot_compras_total:,.2f}", weight="bold", size=11, width=85, text_align=ft.TextAlign.RIGHT),
            ]))
            self.doc_cuerpo.controls.append(ft.Divider(height=15, color="transparent"))

            # 2. SECCIÓN VENTAS
            self.doc_cuerpo.controls.append(ft.Container(content=ft.Text("VENTAS (IVA GENERADO EN SALIDAS)", weight="bold", size=13, color="blue700"), padding=ft.padding.only(top=10, bottom=5)))
            self.doc_cuerpo.controls.append(ft.Row([
                ft.Text("FECHA", weight="bold", size=10, width=65),
                ft.Text("DOC.", weight="bold", size=10, width=80),
                ft.Text("INSUMO", weight="bold", size=10, expand=True),
                ft.Text("BASE", weight="bold", size=10, width=80, text_align=ft.TextAlign.RIGHT),
                ft.Text("IVA", weight="bold", size=10, width=70, text_align=ft.TextAlign.RIGHT),
                ft.Text("TOTAL", weight="bold", size=10, width=85, text_align=ft.TextAlign.RIGHT),
            ]))
            self.doc_cuerpo.controls.append(ft.Divider(height=1, color="black"))

            for v in ventas_filtradas:
                self.doc_cuerpo.controls.append(ft.Row([
                    ft.Text(v['fecha'], size=10, width=65),
                    ft.Text(v['doc'], size=10, width=80, no_wrap=True),
                    ft.Text(v['insumo'], size=10, expand=True, no_wrap=True),
                    ft.Text(f"${v['base']:,.2f}", size=10, width=80, text_align=ft.TextAlign.RIGHT),
                    ft.Text(f"${v['iva']:,.2f}", size=10, width=70, text_align=ft.TextAlign.RIGHT),
                    ft.Text(f"${v['total']:,.2f}", size=10, width=85, text_align=ft.TextAlign.RIGHT),
                ]))

            self.doc_cuerpo.controls.append(ft.Row([
                ft.Text("TOTAL VENTAS:", weight="bold", size=11, expand=True, text_align=ft.TextAlign.RIGHT),
                ft.Text(f"${tot_ventas_base:,.2f}", weight="bold", size=11, width=80, text_align=ft.TextAlign.RIGHT),
                ft.Text(f"${tot_ventas_iva:,.2f}", weight="bold", size=11, width=70, text_align=ft.TextAlign.RIGHT, color="blue700"),
                ft.Text(f"${tot_ventas_total:,.2f}", weight="bold", size=11, width=85, text_align=ft.TextAlign.RIGHT),
            ]))
            self.doc_cuerpo.controls.append(ft.Divider(height=2, color="black"))

            color_bal = "red" if balance_iva > 0 else "green"
            lbl_bal = "BALANCE NETO DE IVA (POR PAGAR)" if balance_iva > 0 else "BALANCE NETO DE IVA (A FAVOR)"
            self.doc_cuerpo.controls.append(
                ft.Row([
                    ft.Text(f"{lbl_bal}:", weight="bold", size=13, expand=True, text_align=ft.TextAlign.RIGHT),
                    ft.Text(f"${balance_iva:,.2f}", weight="bold", size=13, width=120, text_align=ft.TextAlign.RIGHT, color=color_bal),
                ])
            )
            self.doc_cuerpo.controls.append(ft.Divider(height=4, color="black"))

    def _generar_valorizacion(self, detalle, fecha_corte=None):
        raw_inv = self.db._db.get_all("catalogo_insumos?select=*&order=codigo_insumo.asc") or []

        if not raw_inv:
            self.doc_cuerpo.controls.append(
                ft.Container(content=ft.Text("No hay datos para los filtros seleccionados.", size=14, color="grey"), padding=30, alignment=ft.alignment.center)
            )
            self.current_data = {}
            self.current_total = 0
            return

        # Si hay fecha_corte o periodo, calcular existencias acumuladas a esa fecha
        filtro_fecha = f"&fecha=lte.{fecha_corte}T23:59:59" if fecha_corte else ""
        filtro_ajuste = f"&fecha_ajuste=lte.{fecha_corte}T23:59:59" if fecha_corte else ""

        raw_compras = self.db._db.get_all(f"registro_compras?select=codigo_insumo,cantidad&estado_registro=eq.VÁLIDO{filtro_fecha}") or []
        raw_ventas = self.db._db.get_all(f"registro_ventas?select=codigo_insumo,cantidad&estado_registro=neq.ANULADO{filtro_fecha}") or []
        raw_ajustes = self.db._db.get_all(f"registro_ajustes_inventario?select=codigo_insumo,cantidad,tipo_ajuste&estado_registro=eq.VÁLIDO{filtro_ajuste}") or []

        compras_map = {}
        for c in raw_compras:
            cod = str(c.get('codigo_insumo') or '')
            compras_map[cod] = compras_map.get(cod, 0.0) + float(c.get('cantidad') or 0)

        ventas_map = {}
        for v in raw_ventas:
            cod = str(v.get('codigo_insumo') or '')
            ventas_map[cod] = ventas_map.get(cod, 0.0) + float(v.get('cantidad') or 0)

        ajustes_map = {}
        for a in raw_ajustes:
            cod = str(a.get('codigo_insumo') or '')
            es_ent = str(a.get('tipo_ajuste') or '').upper() in ('AJUSTE_ENTRADA', 'ENTRADA_POR_SOBRANTE')
            cant = float(a.get('cantidad') or 0)
            ajustes_map[cod] = ajustes_map.get(cod, 0.0) + (cant if es_ent else -cant)

        agrupacion = {}
        gran_total_costo = 0.0
        gran_total_cant = 0.0

        for item in raw_inv:
            cat = (item.get("categoria") or "SIN CATEGORIA").strip().upper()
            cod = str(item.get("codigo_insumo") or "")
            stk_ini = float(item.get("stock_inicial") or 0)
            ent = compras_map.get(cod, 0.0)
            sal = ventas_map.get(cod, 0.0)
            aj = ajustes_map.get(cod, 0.0)

            stock_real = stk_ini + ent - sal + aj
            costo_u = float(item.get("costo_unitario") or 0)
            
            # REGLA DE NEGOCIO: Un informe de valorización evalúa existencias reales positivas.
            stock_val = max(0.0, stock_real)
            costo_total = stock_val * costo_u

            if stock_val > 0 and costo_total > 0:
                if cat not in agrupacion:
                    agrupacion[cat] = {"items": [], "subtotal": 0.0, "cant_total": 0.0}

                agrupacion[cat]["items"].append({
                    "codigo": cod,
                    "nombre": item.get("nombre"),
                    "stock": stock_val,
                    "costo_u": costo_u,
                    "total": costo_total
                })
                agrupacion[cat]["subtotal"] += costo_total
                agrupacion[cat]["cant_total"] += stock_val
                gran_total_costo += costo_total
                gran_total_cant += stock_val

        # Guardar en memoria para exportación PDF/Excel
        self.current_data = agrupacion
        self.current_total = gran_total_costo
        self.current_periodo = self.doc_header_periodo.value

        if detalle == "Resumido":
            self._dibujar_resumido(agrupacion, "CATEGORÍA", gran_total_costo, gran_total_cant, "GRAN TOTAL VALORIZACIÓN")
        else:
            self.doc_cuerpo.controls.append(
                ft.Row([
                    ft.Text("CÓDIGO", weight="bold", size=11, width=60),
                    ft.Text("INSUMO", weight="bold", size=11, expand=True),
                    ft.Text("CANT.", weight="bold", size=11, width=60, text_align=ft.TextAlign.RIGHT),
                    ft.Text("COSTO U.", weight="bold", size=11, width=80, text_align=ft.TextAlign.RIGHT),
                    ft.Text("TOTAL", weight="bold", size=11, width=100, text_align=ft.TextAlign.RIGHT),
                ])
            )
            self.doc_cuerpo.controls.append(ft.Divider(height=1, color="black"))

            for cat, datos_cat in sorted(agrupacion.items()):
                self.doc_cuerpo.controls.append(
                    ft.Container(content=ft.Text(f"GRUPO: {cat.upper()}", weight="bold", size=12, color=Config.COLOR_PRIMARY), padding=ft.padding.only(top=10, bottom=5))
                )
                for i in sorted(datos_cat["items"], key=lambda x: x["nombre"]):
                    self.doc_cuerpo.controls.append(
                        ft.Row([
                            ft.Text(i['codigo'], size=11, width=60),
                            ft.Text(i['nombre'], size=11, expand=True, no_wrap=True),
                            ft.Text(f"{i['stock']:g}", size=11, width=60, text_align=ft.TextAlign.RIGHT),
                            ft.Text(f"${i['costo_u']:,.2f}", size=11, width=80, text_align=ft.TextAlign.RIGHT),
                            ft.Text(f"${i['total']:,.2f}", size=11, width=100, text_align=ft.TextAlign.RIGHT),
                        ])
                    )
                self.doc_cuerpo.controls.append(
                    ft.Row([
                        ft.Text(f"Total {cat}:", weight="bold", size=11, expand=True, text_align=ft.TextAlign.RIGHT),
                        ft.Text(f"${datos_cat['subtotal']:,.2f}", weight="bold", size=12, width=100, text_align=ft.TextAlign.RIGHT),
                    ])
                )
                self.doc_cuerpo.controls.append(ft.Divider(height=1, color="#eeeeee"))

            self.doc_cuerpo.controls.append(ft.Divider(height=2, color="black"))
            self.doc_cuerpo.controls.append(
                ft.Row([
                    ft.Text("GRAN TOTAL VALORIZACIÓN:", weight="bold", size=14, expand=True, text_align=ft.TextAlign.RIGHT),
                    ft.Text(f"${gran_total_costo:,.2f}", weight="bold", size=14, width=150, text_align=ft.TextAlign.RIGHT),
                ])
            )
            self.doc_cuerpo.controls.append(ft.Divider(height=4, color="black"))

    def _generar_compras(self, fecha_inicio, fecha_fin, detalle):
        data = self.db._db.get_all("registro_compras?select=*,catalogo_insumos(nombre)&estado_registro=eq.VÁLIDO&order=fecha.asc") or []
        agrupacion = {}
        gran_total = 0.0
        gran_total_cant = 0.0

        for item in data:
            fecha = str(item.get("fecha", ""))[:10]
            if not (fecha_inicio <= fecha <= fecha_fin): continue

            proveedor = item.get("proveedor", "Desconocido")
            costo_total = float(item.get("costo_total") or 0)
            cant = float(item.get("cantidad", 0))

            if proveedor not in agrupacion:
                agrupacion[proveedor] = {"items": [], "subtotal": 0.0, "cant_total": 0.0}

            agrupacion[proveedor]["items"].append({
                "fecha": fecha,
                "factura": item.get("numero_factura", "") or item.get("numero_entrada", ""),
                "insumo": item.get("catalogo_insumos", {}).get("nombre", "") if isinstance(item.get("catalogo_insumos"), dict) else item.get("descripcion", ""),
                "cant": cant,
                "total": costo_total
            })
            agrupacion[proveedor]["subtotal"] += costo_total
            agrupacion[proveedor]["cant_total"] += cant
            gran_total += costo_total
            gran_total_cant += cant

        self.current_data = agrupacion
        self.current_total = gran_total
        self.current_periodo = self.doc_header_periodo.value

        if detalle == "Resumido":
            self._dibujar_resumido(agrupacion, "PROVEEDOR", gran_total, gran_total_cant)
        else:
            self._dibujar_tabla_financiera(agrupacion, "PROVEEDOR", gran_total, ["FECHA", "FACTURA", "INSUMO", "CANT.", "TOTAL"])

    def _generar_ventas(self, fecha_inicio, fecha_fin, detalle):
        data = self.db._db.get_all("registro_ventas?select=*,catalogo_insumos(nombre,categoria)&estado_registro=neq.ANULADO&order=fecha.asc") or []
        agrupacion = {}
        gran_total = 0.0
        gran_total_cant = 0.0
        modo_agrup = self.drop_agrupacion_ventas.value or "Tipo de Documento (POS / Remisión)"

        if modo_agrup == "Tipo de Documento (POS / Remisión)":
            label_grupo = "TIPO DE DOCUMENTO"
        elif modo_agrup == "Categoría":
            label_grupo = "CATEGORÍA"
        else:
            label_grupo = "DOCUMENTO Y CATEGORÍA"

        for item in data:
            if item.get("estado_registro") == "ANULADO":
                continue
            fecha = parsear_a_fecha_local(item.get("fecha"))
            if not (fecha_inicio <= fecha <= fecha_fin):
                continue

            tipo_doc = str(item.get("tipo_documento") or "Remisión").strip()
            if "pos" in tipo_doc.lower():
                tipo_doc_fmt = "FACTURA POS"
            elif "remi" in tipo_doc.lower():
                tipo_doc_fmt = "REMISIÓN"
            else:
                tipo_doc_fmt = tipo_doc.upper()

            cat = str(item.get("catalogo_insumos", {}).get("categoria") or "SIN CATEGORIA").strip().upper()

            if modo_agrup == "Tipo de Documento (POS / Remisión)":
                clave_grupo = tipo_doc_fmt
            elif modo_agrup == "Categoría":
                clave_grupo = cat
            else:
                clave_grupo = f"{tipo_doc_fmt} • {cat}"

            if clave_grupo not in agrupacion:
                agrupacion[clave_grupo] = {"items": [], "subtotal": 0.0, "cant_total": 0.0}

            total = float(item.get("total") or 0)
            cant = float(item.get("cantidad", 0))
            factura_num = str(item.get("factura_no") or "S/D")
            doc_display = f"{factura_num} ({tipo_doc_fmt})" if modo_agrup == "Categoría" else factura_num

            agrupacion[clave_grupo]["items"].append({
                "fecha": fecha,
                "factura": doc_display,
                "insumo": item.get("descripcion") or item.get("catalogo_insumos", {}).get("nombre", ""),
                "cant": cant,
                "total": total
            })
            agrupacion[clave_grupo]["subtotal"] += total
            agrupacion[clave_grupo]["cant_total"] += cant
            gran_total += total
            gran_total_cant += cant

        self.current_data = agrupacion
        self.current_total = gran_total
        self.current_periodo = self.doc_header_periodo.value

        if detalle == "Resumido":
            self._dibujar_resumido(agrupacion, label_grupo, gran_total, gran_total_cant)
        else:
            self._dibujar_tabla_financiera(agrupacion, label_grupo, gran_total, ["FECHA", "DOC / FACTURA", "INSUMO", "CANT.", "INGRESOS"])

    def _generar_ajustes(self, fecha_inicio, fecha_fin, detalle):
        data = self.db.get_ajustes_inventario()
        agrupacion = {}
        gran_total_neto = 0.0
        gran_total_cant = 0.0

        from core.fecha_utils import parsear_a_fecha_local
        for item in data:
            if item.get("estado_registro") != "VÁLIDO": continue

            fecha = parsear_a_fecha_local(item.get("fecha_ajuste"))
            if not (fecha_inicio <= fecha <= fecha_fin): continue

            tipo = "ENTRADAS (+)" if item.get("tipo_ajuste") in ('AJUSTE_ENTRADA', 'ENTRADA_POR_SOBRANTE') else "SALIDAS (-)"
            if tipo not in agrupacion:
                agrupacion[tipo] = {"items": [], "subtotal": 0.0, "cant_total": 0.0}

            costo = float(item.get("costo_total_ajuste") or 0)
            cant = float(item.get("cantidad", 0))

            raw_motivo = item.get("motivo", "")
            raw_obs = item.get("observacion", "") or item.get("observaciones", "")
            if not raw_motivo and item.get("motivo_observacion"):
                partes = item.get("motivo_observacion").split("]", 1)
                raw_motivo = partes[0].replace("[", "").strip() if len(partes) > 1 else item.get("motivo_observacion")
                raw_obs = partes[1].strip() if len(partes) > 1 else ""

            agrupacion[tipo]["items"].append({
                "fecha": fecha,
                "motivo": raw_motivo,
                "obs": raw_obs,
                "insumo": item.get("catalogo_insumos", {}).get("nombre", ""),
                "cant": cant,
                "total": costo
            })
            agrupacion[tipo]["subtotal"] += costo
            agrupacion[tipo]["cant_total"] += cant
            gran_total_neto += costo if tipo == "ENTRADAS (+)" else -costo
            gran_total_cant += cant if tipo == "ENTRADAS (+)" else -cant

        self.current_data = agrupacion
        self.current_total = gran_total_neto
        self.current_periodo = self.doc_header_periodo.value

        if detalle == "Resumido":
            self._dibujar_resumido(agrupacion, "TIPO DE AJUSTE", gran_total_neto, gran_total_cant, "IMPACTO NETO")
        else:
            self._dibujar_tabla_financiera(agrupacion, "TIPO DE AJUSTE", gran_total_neto, ["FECHA", "MOTIVO", "OBSERVACIÓN", "INSUMO", "CANT.", "COSTO TOTAL"], label_gran_total="IMPACTO NETO", is_ajuste=True)

    def _dibujar_resumido(self, agrupacion, label_grupo, gran_total_val, gran_total_cant=0, label_gran_total="GRAN TOTAL"):
        self.doc_cuerpo.controls.append(ft.Row([
            ft.Text(label_grupo, weight="bold", size=11, expand=True),
            ft.Text("CANT. TOTAL", weight="bold", size=11, width=100, text_align=ft.TextAlign.RIGHT),
            ft.Text("VALOR TOTAL", weight="bold", size=11, width=120, text_align=ft.TextAlign.RIGHT),
        ]))
        self.doc_cuerpo.controls.append(ft.Divider(height=1, color="black"))

        for grupo, datos in sorted(agrupacion.items()):
            self.doc_cuerpo.controls.append(ft.Row([
                ft.Text(grupo.upper(), size=11, expand=True, weight="bold", color=Config.COLOR_PRIMARY),
                ft.Text(f"{datos.get('cant_total', 0):g}", size=11, width=100, text_align=ft.TextAlign.RIGHT),
                ft.Text(f"${datos['subtotal']:,.2f}", size=11, width=120, text_align=ft.TextAlign.RIGHT),
            ]))
            self.doc_cuerpo.controls.append(ft.Divider(height=1, color="#eeeeee"))

        self.doc_cuerpo.controls.append(ft.Divider(height=2, color="black"))
        color_total = "red" if gran_total_val < 0 else "black"
        self.doc_cuerpo.controls.append(ft.Row([
            ft.Text(f"{label_gran_total}:", weight="bold", size=14, expand=True, text_align=ft.TextAlign.RIGHT),
            ft.Text(f"{gran_total_cant:g}", weight="bold", size=14, width=100, text_align=ft.TextAlign.RIGHT),
            ft.Text(f"${gran_total_val:,.2f}", weight="bold", size=14, width=120, text_align=ft.TextAlign.RIGHT, color=color_total),
        ]))

    def _dibujar_tabla_financiera(self, agrupacion, label_grupo, gran_total, headers, label_gran_total="GRAN TOTAL", is_ajuste=False):
        if not agrupacion:
            self.doc_cuerpo.controls.append(ft.Container(content=ft.Text("No hay datos para los filtros seleccionados.", size=14, color="grey"), padding=30, alignment=ft.alignment.center))
            return

        # Cabeceras
        if is_ajuste:
            self.doc_cuerpo.controls.append(ft.Row([
                ft.Text(headers[0], weight="bold", size=11, width=65),
                ft.Text(headers[1], weight="bold", size=11, width=90),
                ft.Text(headers[2], weight="bold", size=11, expand=True),
                ft.Text(headers[3], weight="bold", size=11, width=100),
                ft.Text(headers[4], weight="bold", size=11, width=45, text_align=ft.TextAlign.RIGHT),
                ft.Text(headers[5], weight="bold", size=11, width=70, text_align=ft.TextAlign.RIGHT),
            ]))
        else:
            self.doc_cuerpo.controls.append(ft.Row([
                ft.Text(headers[0], weight="bold", size=11, width=70),
                ft.Text(headers[1], weight="bold", size=11, width=90),
                ft.Text(headers[2], weight="bold", size=11, expand=True),
                ft.Text(headers[3], weight="bold", size=11, width=60, text_align=ft.TextAlign.RIGHT),
                ft.Text(headers[4], weight="bold", size=11, width=100, text_align=ft.TextAlign.RIGHT),
            ]))
        self.doc_cuerpo.controls.append(ft.Divider(height=1, color="black"))

        for grupo, datos in sorted(agrupacion.items()):
            self.doc_cuerpo.controls.append(ft.Container(content=ft.Text(f"{label_grupo}: {grupo.upper()}", weight="bold", size=12, color=Config.COLOR_PRIMARY), padding=ft.padding.only(top=10, bottom=5)))
            for i in datos["items"]:
                if is_ajuste:
                    fila_ui = ft.Row([
                        ft.Text(i['fecha'], size=11, width=65),
                        ft.Text(i['motivo'], size=11, width=90, no_wrap=True, weight="bold"),
                        ft.Text(i['obs'], size=11, expand=True), # Observación toma el espacio restante
                        ft.Text(i['insumo'], size=11, width=100, no_wrap=True),
                        ft.Text(f"{i['cant']:g}", size=11, width=45, text_align=ft.TextAlign.RIGHT),
                        ft.Text(f"${i['total']:,.2f}", size=11, width=70, text_align=ft.TextAlign.RIGHT),
                    ])
                else:
                    fila_ui = ft.Row([
                        ft.Text(i['fecha'], size=11, width=70),
                        ft.Text(i['factura'], size=11, width=90, no_wrap=True),
                        ft.Text(i['insumo'], size=11, expand=True, no_wrap=True),
                        ft.Text(f"{i['cant']:g}", size=11, width=60, text_align=ft.TextAlign.RIGHT),
                        ft.Text(f"${i['total']:,.2f}", size=11, width=100, text_align=ft.TextAlign.RIGHT),
                    ])
                self.doc_cuerpo.controls.append(fila_ui)
            self.doc_cuerpo.controls.append(ft.Row([
                ft.Text(f"Total {grupo}:", weight="bold", size=11, expand=True, text_align=ft.TextAlign.RIGHT),
                ft.Text(f"${datos['subtotal']:,.2f}", weight="bold", size=12, width=100, text_align=ft.TextAlign.RIGHT),
            ]))
            self.doc_cuerpo.controls.append(ft.Divider(height=1, color="#eeeeee"))

        self.doc_cuerpo.controls.append(ft.Divider(height=2, color="black"))
        color_total = "red" if gran_total < 0 else "black"
        self.doc_cuerpo.controls.append(ft.Row([
            ft.Text(f"{label_gran_total}:", weight="bold", size=14, expand=True, text_align=ft.TextAlign.RIGHT),
            ft.Text(f"${gran_total:,.2f}", weight="bold", size=14, width=150, text_align=ft.TextAlign.RIGHT, color=color_total),
        ]))
        self.doc_cuerpo.controls.append(ft.Divider(height=4, color="black"))

    def _generar_recaudos(self, fecha_inicio, fecha_fin, detalle):
        pagos = self.db.cartera_repo._get_todos_pagos_deduplicados()
        detalles_pagos = self.db.cartera_repo._get_detalles_deduplicados()

        # Mapear facturas afectadas por cada id_pago
        facturas_por_pago = {}
        for d in detalles_pagos:
            p_id = d.get("id_pago")
            f_no = str(d.get("factura_no", "")).strip()
            if p_id and f_no:
                facturas_por_pago.setdefault(p_id, []).append(f_no)

        # Mapear vendedores de clientes
        clientes_db = self.db.clientes_repo.get_clientes()
        vendedor_map = {}
        for c in clientes_db:
            nom_c = self.db.clientes_repo.normalizar_nombre_cliente(c.get("nombre"))
            vendedor_map[nom_c] = c.get("vendedor_encargado") or "Sin Asignar"

        # Agrupar pagos válidos en el rango de fechas
        agrupacion = {}
        gran_total = 0.0
        tot_efectivo = 0.0
        tot_transf = 0.0
        cant_pagos_total = 0

        for p in pagos:
            f_pago = str(p.get("fecha_pago") or "")[:10]
            if not (fecha_inicio <= f_pago <= fecha_fin):
                continue

            cli_nom = self.db.clientes_repo.normalizar_nombre_cliente(p.get("nombre_cliente"))
            monto = float(p.get("monto_total") or 0.0)
            metodo = str(p.get("metodo_pago") or "EFECTIVO").upper()
            p_id = p.get("id_pago")
            facs_list = facturas_por_pago.get(p_id, [])
            facs_str = ", ".join([f"#{f}" for f in facs_list]) if facs_list else "Global (FIFO)"
            banco_ref = f"{p.get('banco_origen') or ''} {p.get('referencia_comprobante') or ''}".strip() or "-"
            vend = vendedor_map.get(cli_nom, "Sin Asignar")

            if cli_nom not in agrupacion:
                agrupacion[cli_nom] = {
                    "vendedor": vend,
                    "items": [],
                    "subtotal": 0.0,
                    "cant_total": 0
                }

            agrupacion[cli_nom]["items"].append({
                "fecha": f_pago,
                "metodo": metodo,
                "banco_ref": banco_ref,
                "facturas": facs_str,
                "monto": monto,
                "usuario": p.get("usuario_registro") or "admin"
            })
            agrupacion[cli_nom]["subtotal"] += monto
            agrupacion[cli_nom]["cant_total"] += 1

            gran_total += monto
            if "TRANSFERENCIA" in metodo:
                tot_transf += monto
            else:
                tot_efectivo += monto
            cant_pagos_total += 1

        self.current_data = agrupacion
        self.current_total = gran_total
        self.current_periodo = self.doc_header_periodo.value

        if not agrupacion:
            self.doc_cuerpo.controls.append(
                ft.Container(
                    content=ft.Text("No se encontraron recaudos registrados en el periodo seleccionado.", size=13, color="grey", italic=True),
                    alignment=ft.alignment.center,
                    padding=30
                )
            )
            return

        if detalle == "Resumido":
            # Tabla Resumida de Recaudos por Cliente
            self.doc_cuerpo.controls.append(ft.Row([
                ft.Text("CLIENTE", weight="bold", size=11, expand=True),
                ft.Text("ENCARGADO / VENDEDOR", weight="bold", size=11, width=180),
                ft.Text("CANT. PAGOS", weight="bold", size=11, width=100, text_align=ft.TextAlign.RIGHT),
                ft.Text("TOTAL RECAUDADO", weight="bold", size=11, width=130, text_align=ft.TextAlign.RIGHT),
            ]))
            self.doc_cuerpo.controls.append(ft.Divider(height=1, color="black"))

            for cli_nom, datos in sorted(agrupacion.items()):
                self.doc_cuerpo.controls.append(ft.Row([
                    ft.Text(cli_nom, size=11, weight="bold", expand=True, color=Config.COLOR_PRIMARY),
                    ft.Text(datos["vendedor"], size=11, width=180, color="purple800" if datos["vendedor"] != "Sin Asignar" else "grey600"),
                    ft.Text(f"{datos['cant_total']}", size=11, width=100, text_align=ft.TextAlign.RIGHT),
                    ft.Text(f"${datos['subtotal']:,.2f}", size=11, weight="bold", width=130, text_align=ft.TextAlign.RIGHT, color="green800"),
                ]))
                self.doc_cuerpo.controls.append(ft.Divider(height=1, color="#eeeeee"))

            self.doc_cuerpo.controls.append(ft.Divider(height=2, color="black"))
            self.doc_cuerpo.controls.append(ft.Row([
                ft.Text("GRAN TOTAL RECAUDADO:", weight="bold", size=14, expand=True, text_align=ft.TextAlign.RIGHT),
                ft.Text(f"${gran_total:,.2f}", weight="bold", size=14, width=150, text_align=ft.TextAlign.RIGHT, color="green800"),
            ]))
            self.doc_cuerpo.controls.append(ft.Row([
                ft.Text(f"Total Efectivo: ${tot_efectivo:,.2f}  •  Total Transferencias: ${tot_transf:,.2f}", size=10.5, color=Config.COLOR_TEXT_MUTED, expand=True, text_align=ft.TextAlign.RIGHT)
            ]))
        else:
            # Tabla Detallada por Cliente
            self.doc_cuerpo.controls.append(ft.Row([
                ft.Text("FECHA", weight="bold", size=11, width=80),
                ft.Text("MÉTODO", weight="bold", size=11, width=95),
                ft.Text("BANCO / COMPROBANTE", weight="bold", size=11, width=130),
                ft.Text("FACTURAS APLICADAS", weight="bold", size=11, expand=True),
                ft.Text("MONTO RECAUDADO", weight="bold", size=11, width=120, text_align=ft.TextAlign.RIGHT),
            ]))
            self.doc_cuerpo.controls.append(ft.Divider(height=1, color="black"))

            for cli_nom, datos in sorted(agrupacion.items()):
                # Banner del Cliente con Encargado
                self.doc_cuerpo.controls.append(
                    ft.Container(
                        content=ft.Row([
                            ft.Text(f"CLIENTE: {cli_nom}", size=12, weight="bold", color=Config.COLOR_PRIMARY, expand=True),
                            ft.Container(
                                content=ft.Row([
                                    ft.Icon(ft.icons.BADGE_OUTLINED, size=12, color="purple800"),
                                    ft.Text(f"Encargado: {datos['vendedor']}", size=10, weight="bold", color="purple900")
                                ], spacing=3, tight=True),
                                bgcolor="#F5F3FF",
                                padding=ft.padding.symmetric(horizontal=6, vertical=2),
                                border_radius=4
                            )
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                        bgcolor="#F1F5F9",
                        padding=ft.padding.symmetric(horizontal=8, vertical=4),
                        border_radius=4,
                        margin=ft.margin.only(top=8, bottom=2)
                    )
                )

                for item in datos["items"]:
                    self.doc_cuerpo.controls.append(
                        ft.Row([
                            ft.Text(item["fecha"], size=10.5, width=80),
                            ft.Text(item["metodo"], size=10.5, width=95),
                            ft.Text(item["banco_ref"][:22], size=10.5, width=130, no_wrap=True),
                            ft.Text(item["facturas"], size=10.5, expand=True, no_wrap=True),
                            ft.Text(f"${item['monto']:,.2f}", size=10.5, weight="bold", width=120, text_align=ft.TextAlign.RIGHT, color="green800"),
                        ])
                    )

                self.doc_cuerpo.controls.append(
                    ft.Row([
                        ft.Text(f"Subtotal Recaudado {cli_nom}:", weight="bold", size=11, expand=True, text_align=ft.TextAlign.RIGHT),
                        ft.Text(f"${datos['subtotal']:,.2f}", weight="bold", size=11.5, width=120, text_align=ft.TextAlign.RIGHT, color="green800"),
                    ])
                )
                self.doc_cuerpo.controls.append(ft.Divider(height=1, color="#eeeeee"))

            self.doc_cuerpo.controls.append(ft.Divider(height=2, color="black"))
            self.doc_cuerpo.controls.append(ft.Row([
                ft.Text("GRAN TOTAL RECAUDADO:", weight="bold", size=14, expand=True, text_align=ft.TextAlign.RIGHT),
                ft.Text(f"${gran_total:,.2f}", weight="bold", size=14, width=150, text_align=ft.TextAlign.RIGHT, color="green800"),
            ]))
            self.doc_cuerpo.controls.append(ft.Row([
                ft.Text(f"Total Efectivo: ${tot_efectivo:,.2f}  •  Total Transferencias: ${tot_transf:,.2f}", size=11, color=Config.COLOR_TEXT_MUTED, expand=True, text_align=ft.TextAlign.RIGHT)
            ]))
            self.doc_cuerpo.controls.append(ft.Divider(height=4, color="black"))

    def _generar_kpis(self, fecha_inicio, fecha_fin, mes_periodo=None):
        fecha_corte = fecha_fin if fecha_fin < "2100-01-01" else None
        res_ven = self.db.get_ventas_summary(fecha_corte=fecha_corte)
        res_com = self.db.get_compras_summary(fecha_corte=fecha_corte)
        kpis_inv = self.db.get_inventario_kpis(fecha_corte=fecha_corte)

        val_inv = float(kpis_inv.get('valor_inventario', 0) or 0)
        ingresos = float(res_ven.get('total_mes') or res_ven.get('total_historico') or 0)
        compras = float(res_com.get('total_mes') or res_com.get('total_historico') or 0)

        rentabilidad = ((ingresos - compras) / ingresos) * 100 if ingresos > 0 else 0
        rotacion = ingresos / val_inv if val_inv > 0 else 0

        # Cartera y Recaudos
        kpis_cartera, _, _ = self.db.cartera_repo.get_resumen_cartera(mes_periodo=mes_periodo)
        tot_cartera_fact = float(kpis_cartera.get("total_ventas") or 0.0)
        tot_cartera_rec = float(kpis_cartera.get("total_recaudado") or 0.0)
        tot_efectivo = float(kpis_cartera.get("total_efectivo") or 0.0)
        tot_transf = float(kpis_cartera.get("total_transferencias") or 0.0)
        tot_saldo_pend = float(kpis_cartera.get("total_saldo_pendiente") or 0.0)
        clientes_con_deuda = int(kpis_cartera.get("clientes_con_deuda") or 0)
        efectividad_cobro = (tot_cartera_rec / tot_cartera_fact * 100) if tot_cartera_fact > 0 else 0.0

        self.current_data = {
            "tipo": "KPIS",
            "val_inv": val_inv,
            "compras": compras,
            "ingresos": ingresos,
            "iva_mes": float(res_ven.get('iva_mes', 0) or 0),
            "rentabilidad": rentabilidad,
            "rotacion": rotacion,
            "tot_cartera_fact": tot_cartera_fact,
            "tot_cartera_rec": tot_cartera_rec,
            "tot_efectivo": tot_efectivo,
            "tot_transf": tot_transf,
            "tot_saldo_pend": tot_saldo_pend,
            "clientes_con_deuda": clientes_con_deuda,
            "efectividad_cobro": efectividad_cobro
        }
        self.current_total = tot_cartera_rec
        self.current_periodo = self.doc_header_periodo.value

        def _crear_kpi_fila(label, valor, color="black"):
            return ft.Row([
                ft.Text(label, size=13, expand=True),
                ft.Text(valor, size=14, weight="bold", color=color, width=150, text_align=ft.TextAlign.RIGHT)
            ])

        self.doc_cuerpo.controls.extend([
            ft.Container(content=ft.Text("NOTA: Este resumen consolida las métricas operativas y financieras del inventario, ventas y la gestión de recaudo de cartera para el periodo seleccionado.", size=10, color="orange", italic=True), padding=10, bgcolor="#fff3cd", border_radius=5),
            ft.Divider(height=10, color="transparent"),
            ft.Text("MÉTRICAS DE INVENTARIO Y COSTOS", weight="bold", size=14, color=Config.COLOR_PRIMARY),
            ft.Divider(height=1, color="black"),
            _crear_kpi_fila("Valorización Actual del Inventario", f"${val_inv:,.2f}"),
            _crear_kpi_fila("Total Compras (Periodo)", f"${compras:,.2f}"),
            ft.Divider(height=20, color="transparent"),
            ft.Text("MÉTRICAS DE VENTAS E INGRESOS", weight="bold", size=14, color=Config.COLOR_PRIMARY),
            ft.Divider(height=1, color="black"),
            _crear_kpi_fila("Total Ventas (Periodo)", f"${ingresos:,.2f}", "green"),
            _crear_kpi_fila("IVA Recaudado (Periodo)", f"${float(res_ven.get('iva_mes', 0) or 0):,.2f}"),
            ft.Divider(height=20, color="transparent"),
            ft.Text("MÉTRICAS DE RECAUDO Y CARTERA", weight="bold", size=14, color=Config.COLOR_PRIMARY),
            ft.Divider(height=1, color="black"),
            _crear_kpi_fila("Total Facturado Cartera", f"${tot_cartera_fact:,.2f}"),
            _crear_kpi_fila("Total Recaudado (Cobrado)", f"${tot_cartera_rec:,.2f}", "green"),
            _crear_kpi_fila("  ↳ Recaudo en Efectivo", f"${tot_efectivo:,.2f}"),
            _crear_kpi_fila("  ↳ Recaudo en Transferencias", f"${tot_transf:,.2f}"),
            _crear_kpi_fila("Saldo Pendiente por Cobrar", f"${tot_saldo_pend:,.2f}", "red"),
            _crear_kpi_fila("Clientes con Deuda Activa", f"{clientes_con_deuda} clientes"),
            _crear_kpi_fila("Efectividad de Recaudo", f"{efectividad_cobro:.1f}%", "green" if efectividad_cobro >= 50 else "orange"),
            ft.Divider(height=20, color="transparent"),
            ft.Text("RENDIMIENTO FINANCIERO", weight="bold", size=14, color=Config.COLOR_PRIMARY),
            ft.Divider(height=1, color="black"),
            _crear_kpi_fila("Margen de Rentabilidad Bruta", f"{rentabilidad:.1f}%", "green" if rentabilidad >= 0 else "red"),
            _crear_kpi_fila("Índice de Rotación", f"{rotacion:.2f}x"),
            ft.Divider(height=4, color="black")
        ])

    def exportar_pdf(self, e):
        if not hasattr(self, 'current_data') or not self.current_data:
            self.page.snack_bar = ft.SnackBar(ft.Text("Primero genera la previsualización del informe."), bgcolor="orange")
            self.page.snack_bar.open = True
            self.page.update()
            return

        # Garantizar registración en overlay antes de invocar el diálogo
        if self.page:
            if self.save_pdf_picker not in self.page.overlay:
                self.page.overlay.append(self.save_pdf_picker)
                self.page.update()

        p_val = str(self.drop_filtro_fecha.value or "HISTORICO").replace(" ", "_")
        nombre_sugerido = f"{self.drop_tipo_informe.value.replace(' ', '_')}_{p_val}_{datetime.date.today().strftime('%Y%m%d')}.pdf"
        self.save_pdf_picker.save_file(
            dialog_title="Guardar Informe PDF",
            file_name=nombre_sugerido,
            allowed_extensions=["pdf"]
        )

    def _save_pdf_result(self, e: ft.FilePickerResultEvent):
        if not e.path:
            return
            
        try:
            from fpdf import FPDF
            
            class PDFReport(FPDF):
                def header(instance):
                    instance.set_font("Arial", 'B', 12)
                    instance.cell(0, 6, "TIENDA Y ABARROTES LOS DESECHABLES DE DOÑA MARY SAS", ln=True, align='C')
                    instance.set_font("Arial", 'B', 10)
                    instance.cell(0, 5, "REPORTE OFICIAL DE INVENTARIOS Y OPERACIONES", ln=True, align='C')
                    instance.set_font("Arial", '', 8)
                    instance.cell(0, 4, f"Generado el: {datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')}", ln=True, align='C')
                    instance.ln(4)
                    instance.line(10, instance.get_y(), 200, instance.get_y())
                    instance.ln(4)

                def footer(instance):
                    instance.set_y(-15)
                    instance.set_font("Arial", 'I', 8)
                    instance.cell(0, 10, f"Página {instance.page_no()}/{{nb}}", align='C')

            pdf = PDFReport()
            pdf.alias_nb_pages()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)

            # Título del Informe
            pdf.set_font("Arial", 'B', 11)
            pdf.cell(0, 6, str(self.doc_header_titulo.value).encode('latin-1', 'replace').decode('latin-1'), ln=True)
            pdf.set_font("Arial", '', 9)
            pdf.cell(0, 5, str(self.doc_header_periodo.value).encode('latin-1', 'replace').decode('latin-1'), ln=True)
            pdf.ln(4)

            es_resumido = self.opcion_detalle.value == "Resumido"
            es_ajuste = self.drop_tipo_informe.value == "Historial de Ajustes"
            es_recaudos = self.drop_tipo_informe.value == "Informe de Recaudos"
            es_kpis = self.drop_tipo_informe.value == "Resumen de KPIs"

            if es_kpis:
                # PDF de Resumen de KPIs
                k_data = self.current_data if isinstance(self.current_data, dict) and self.current_data.get("tipo") == "KPIS" else {}
                pdf.set_font("Arial", 'B', 10)
                pdf.cell(0, 6, "1. METRICAS DE INVENTARIO Y COSTOS", ln=True)
                pdf.set_font("Arial", '', 9)
                pdf.cell(130, 6, "Valorizacion Actual del Inventario:", border="B")
                pdf.cell(50, 6, f"${k_data.get('val_inv', 0):,.2f}", border="B", align='R', ln=True)
                pdf.cell(130, 6, "Total Compras (Mes Actual):", border="B")
                pdf.cell(50, 6, f"${k_data.get('compras', 0):,.2f}", border="B", align='R', ln=True)
                pdf.ln(3)

                pdf.set_font("Arial", 'B', 10)
                pdf.cell(0, 6, "2. METRICAS DE VENTAS E INGRESOS", ln=True)
                pdf.set_font("Arial", '', 9)
                pdf.cell(130, 6, "Total Ventas (Mes Actual):", border="B")
                pdf.cell(50, 6, f"${k_data.get('ingresos', 0):,.2f}", border="B", align='R', ln=True)
                pdf.cell(130, 6, "IVA Recaudado (Mes Actual):", border="B")
                pdf.cell(50, 6, f"${k_data.get('iva_mes', 0):,.2f}", border="B", align='R', ln=True)
                pdf.ln(3)

                pdf.set_font("Arial", 'B', 10)
                pdf.cell(0, 6, "3. METRICAS DE RECAUDO Y GESTION DE CARTERA", ln=True)
                pdf.set_font("Arial", '', 9)
                pdf.cell(130, 6, "Total Facturado Cartera:", border="B")
                pdf.cell(50, 6, f"${k_data.get('tot_cartera_fact', 0):,.2f}", border="B", align='R', ln=True)
                pdf.cell(130, 6, "Total Recaudado (Cobrado):", border="B")
                pdf.cell(50, 6, f"${k_data.get('tot_cartera_rec', 0):,.2f}", border="B", align='R', ln=True)
                pdf.cell(130, 6, "  - Recaudo en Efectivo:", border="B")
                pdf.cell(50, 6, f"${k_data.get('tot_efectivo', 0):,.2f}", border="B", align='R', ln=True)
                pdf.cell(130, 6, "  - Recaudo en Transferencias:", border="B")
                pdf.cell(50, 6, f"${k_data.get('tot_transf', 0):,.2f}", border="B", align='R', ln=True)
                pdf.cell(130, 6, "Saldo Pendiente por Cobrar (Cartera Activa):", border="B")
                pdf.cell(50, 6, f"${k_data.get('tot_saldo_pend', 0):,.2f}", border="B", align='R', ln=True)
                pdf.cell(130, 6, "Clientes con Deuda Activa:", border="B")
                pdf.cell(50, 6, f"{k_data.get('clientes_con_deuda', 0)} clientes", border="B", align='R', ln=True)
                pdf.cell(130, 6, "Porcentaje de Efectividad de Recaudo:", border="B")
                pdf.cell(50, 6, f"{k_data.get('efectividad_cobro', 0):.1f}%", border="B", align='R', ln=True)
                pdf.ln(3)

                pdf.set_font("Arial", 'B', 10)
                pdf.cell(0, 6, "4. RENDIMIENTO FINANCIERO", ln=True)
                pdf.set_font("Arial", '', 9)
                pdf.cell(130, 6, "Margen de Rentabilidad Bruta:", border="B")
                pdf.cell(50, 6, f"{k_data.get('rentabilidad', 0):.1f}%", border="B", align='R', ln=True)
                pdf.cell(130, 6, "Indice de Rotacion:", border="B")
                pdf.cell(50, 6, f"{k_data.get('rotacion', 0):.2f}x", border="B", align='R', ln=True)

            elif es_recaudos:
                if es_resumido:
                    pdf.set_font("Arial", 'B', 8)
                    pdf.cell(75, 6, "CLIENTE", border=1)
                    pdf.cell(45, 6, "ENCARGADO / VENDEDOR", border=1)
                    pdf.cell(25, 6, "CANT. PAGOS", border=1, align='R')
                    pdf.cell(45, 6, "TOTAL RECAUDADO", border=1, align='R')
                    pdf.ln()

                    pdf.set_font("Arial", '', 8)
                    for cli_nom, datos in sorted(self.current_data.items()):
                        c_name = str(cli_nom)[:32].encode('latin-1', 'replace').decode('latin-1')
                        v_name = str(datos.get("vendedor") or "Sin Asignar")[:20].encode('latin-1', 'replace').decode('latin-1')
                        pdf.cell(75, 6, c_name, border="L,R,B")
                        pdf.cell(45, 6, v_name, border="L,R,B")
                        pdf.cell(25, 6, f"{datos.get('cant_total', 0)}", border="L,R,B", align='R')
                        pdf.cell(45, 6, f"${datos['subtotal']:,.2f}", border="L,R,B", align='R', ln=True)

                    pdf.ln(4)
                    pdf.set_font("Arial", 'B', 9)
                    pdf.cell(145, 7, "GRAN TOTAL RECAUDADO:", align='R')
                    pdf.cell(45, 7, f"${self.current_total:,.2f}", border=1, align='R', ln=True)
                else:
                    for cli_nom, datos in sorted(self.current_data.items()):
                        c_name = str(cli_nom)[:35].encode('latin-1', 'replace').decode('latin-1')
                        v_name = str(datos.get("vendedor") or "Sin Asignar")[:25].encode('latin-1', 'replace').decode('latin-1')

                        pdf.set_font("Arial", 'B', 8)
                        pdf.cell(0, 6, f"  CLIENTE: {c_name}  |  ENCARGADO: {v_name}", border=1, ln=True)

                        pdf.set_font("Arial", 'B', 7.5)
                        pdf.cell(25, 5, "FECHA", border=1)
                        pdf.cell(30, 5, "METODO", border=1)
                        pdf.cell(45, 5, "BANCO / REF", border=1)
                        pdf.cell(55, 5, "FACTURAS APLICADAS", border=1)
                        pdf.cell(35, 5, "MONTO RECAUDADO", border=1, align='R')
                        pdf.ln()

                        pdf.set_font("Arial", '', 7)
                        for item in datos["items"]:
                            b_ref = str(item.get("banco_ref") or "-")[:22].encode('latin-1', 'replace').decode('latin-1')
                            facs = str(item.get("facturas") or "-")[:30].encode('latin-1', 'replace').decode('latin-1')
                            pdf.cell(25, 5, str(item.get("fecha", "")), border="L")
                            pdf.cell(30, 5, str(item.get("metodo", "EFECTIVO")))
                            pdf.cell(45, 5, b_ref)
                            pdf.cell(55, 5, facs)
                            pdf.cell(35, 5, f"${item.get('monto', 0):,.2f}", border="R", align='R', ln=True)

                        pdf.set_font("Arial", 'B', 8)
                        pdf.cell(155, 5, f"Subtotal Recaudado {c_name}:", border="L,B", align='R')
                        pdf.cell(35, 5, f"${datos['subtotal']:,.2f}", border="R,B", align='R', ln=True)
                        pdf.ln(2)

                    pdf.ln(4)
                    pdf.set_font("Arial", 'B', 9)
                    pdf.cell(155, 7, "GRAN TOTAL RECAUDADO:", align='R')
                    pdf.cell(35, 7, f"${self.current_total:,.2f}", border=1, align='R', ln=True)

            elif es_resumido:
                # Tabla Resumida
                pdf.set_font("Arial", 'B', 8)
                pdf.cell(110, 6, "GRUPO / CATEGORIA", border=1)
                pdf.cell(35, 6, "CANT. TOTAL", border=1, align='R')
                pdf.cell(45, 6, "VALOR TOTAL", border=1, align='R')
                pdf.ln()

                pdf.set_font("Arial", '', 8)
                for grupo, datos in sorted(self.current_data.items()):
                    g_nombre = str(grupo).upper().encode('latin-1', 'replace').decode('latin-1')
                    pdf.cell(110, 6, g_nombre, border="L,R,B")
                    pdf.cell(35, 6, f"{datos.get('cant_total', 0):g}", border="L,R,B", align='R')
                    pdf.cell(45, 6, f"${datos['subtotal']:,.2f}", border="L,R,B", align='R', ln=True)

                pdf.ln(4)
                pdf.set_font("Arial", 'B', 9)
                pdf.cell(145, 7, "TOTAL GENERAL:", align='R')
                pdf.cell(45, 7, f"${self.current_total:,.2f}", border=1, align='R', ln=True)

            else:
                # Tabla Completa
                pdf.set_font("Arial", 'B', 8)
                if es_ajuste:
                    pdf.cell(20, 6, "FECHA", border=1)
                    pdf.cell(30, 6, "MOTIVO", border=1)
                    pdf.cell(50, 6, "OBSERVACION", border=1)
                    pdf.cell(45, 6, "INSUMO", border=1)
                    pdf.cell(20, 6, "CANT.", border=1, align='R')
                    pdf.cell(25, 6, "COSTO TOT.", border=1, align='R')
                else:
                    pdf.cell(20, 6, "FECHA/COD", border=1)
                    pdf.cell(30, 6, "DOC/FACT", border=1)
                    pdf.cell(75, 6, "INSUMO / DESCRIPCION", border=1)
                    pdf.cell(20, 6, "CANT.", border=1, align='R')
                    pdf.cell(20, 6, "COSTO U.", border=1, align='R')
                    pdf.cell(25, 6, "TOTAL", border=1, align='R')
                pdf.ln()

                for grupo, datos in sorted(self.current_data.items()):
                    pdf.set_font("Arial", 'B', 8)
                    pdf.cell(0, 6, f"  GRUPO: {str(grupo).upper().encode('latin-1', 'replace').decode('latin-1')}", border="L,R,B", ln=True)
                    pdf.set_font("Arial", '', 7)

                    for i in datos["items"]:
                        insumo_txt = str(i.get('insumo') or i.get('nombre', '')).encode('latin-1', 'replace').decode('latin-1')[:35]
                        
                        if es_ajuste:
                            motivo_txt = str(i.get('motivo', '')).encode('latin-1', 'replace').decode('latin-1')[:18]
                            obs_txt = str(i.get('obs', '')).encode('latin-1', 'replace').decode('latin-1')[:30]
                            pdf.cell(20, 5, str(i.get('fecha', '')), border="L")
                            pdf.cell(30, 5, motivo_txt)
                            pdf.cell(50, 5, obs_txt)
                            pdf.cell(45, 5, insumo_txt)
                            pdf.cell(20, 5, f"{i.get('cant', 0):g}", align='R')
                            pdf.cell(25, 5, f"${i.get('total', 0):,.2f}", border="R", align='R', ln=True)
                        else:
                            c_code = str(i.get('codigo') or i.get('fecha', ''))
                            c_doc = str(i.get('factura', ''))[:15]
                            c_u = f"${i.get('costo_u', 0):,.2f}" if 'costo_u' in i else "-"
                            pdf.cell(20, 5, c_code, border="L")
                            pdf.cell(30, 5, c_doc)
                            pdf.cell(75, 5, insumo_txt)
                            pdf.cell(20, 5, f"{i.get('stock', i.get('cant', 0)):g}", align='R')
                            pdf.cell(20, 5, c_u, align='R')
                            pdf.cell(25, 5, f"${i.get('total', 0):,.2f}", border="R", align='R', ln=True)

                    pdf.set_font("Arial", 'B', 8)
                    pdf.cell(165, 5, f"Subtotal {grupo}:", border="L,B", align='R')
                    pdf.cell(25, 5, f"${datos['subtotal']:,.2f}", border="R,B", align='R', ln=True)

                pdf.ln(4)
                pdf.set_font("Arial", 'B', 9)
                pdf.cell(165, 7, "TOTAL GENERAL:", align='R')
                pdf.cell(25, 7, f"${self.current_total:,.2f}", border=1, align='R', ln=True)

            pdf.output(e.path)
            if self.page:
                self.page.snack_bar = ft.SnackBar(ft.Text("¡PDF exportado con éxito!"), bgcolor="green")
                self.page.snack_bar.open = True
                self.page.update()

        except Exception as ex:
            if self.page:
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Error al generar PDF: {ex}"), bgcolor="red")
                self.page.snack_bar.open = True
                self.page.update()

    def exportar_excel(self, e):
        # Garantizar registración en overlay antes de invocar el diálogo
        if self.page:
            if self.save_excel_picker not in self.page.overlay:
                self.page.overlay.append(self.save_excel_picker)
                self.page.update()

        nombre_sugerido = f"Inventario_Consolidado_Dona_Mary_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        self.save_excel_picker.save_file(
            dialog_title="Guardar Consolidado Excel",
            file_name=nombre_sugerido,
            allowed_extensions=["xlsx"]
        )

    def _save_excel_result(self, e: ft.FilePickerResultEvent):
        if not e.path:
            return
            
        self.page.snack_bar = ft.SnackBar(ft.Text("Generando consolidado Excel en segundo plano..."), bgcolor="blue")
        self.page.snack_bar.open = True
        self.page.update()
        
        threading.Thread(target=self._worker_generar_excel, args=(e.path,), daemon=True).start()

    def _worker_generar_excel(self, file_path):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from concurrent.futures import ThreadPoolExecutor

            # 1. Descargar todas las fuentes de datos en paralelo para máxima velocidad
            def fetch_inv():
                return self.db._db.get_all("catalogo_insumos?select=*&order=codigo_insumo.asc") or []

            def fetch_compras():
                return self.db._db.get_all("registro_compras?select=*,catalogo_insumos(nombre)&estado_registro=eq.VÁLIDO&order=fecha.asc") or []

            def fetch_ventas():
                return self.db._db.get_all("registro_ventas?select=*,catalogo_insumos(nombre)&estado_registro=neq.ANULADO&order=fecha.asc") or []

            def fetch_ajustes():
                return self.db._db.get_all("registro_ajustes_inventario?select=*,catalogo_insumos(nombre)&estado_registro=eq.VÁLIDO&order=fecha_ajuste.asc") or []

            def fetch_cartera():
                return self.db.cartera_repo.get_resumen_cartera()

            def fetch_pagos():
                return self.db.cartera_repo._get_todos_pagos_deduplicados(), self.db.cartera_repo._get_detalles_deduplicados()

            with ThreadPoolExecutor(max_workers=6) as executor:
                f_inv = executor.submit(fetch_inv)
                f_compras = executor.submit(fetch_compras)
                f_ventas = executor.submit(fetch_ventas)
                f_ajustes = executor.submit(fetch_ajustes)
                f_cartera = executor.submit(fetch_cartera)
                f_pagos = executor.submit(fetch_pagos)

                raw_inv = f_inv.result()
                raw_compras = f_compras.result()
                raw_ventas = f_ventas.result()
                raw_ajustes = f_ajustes.result()
                kpis_cartera, clientes_cartera, docs_cartera = f_cartera.result()
                pagos_cartera, detalles_pagos = f_pagos.result()

            # Mapeo rápido de clientes y pagos
            vendedor_map = {c.get("nombre"): c.get("vendedor_encargado") or "Sin Asignar" for c in clientes_cartera}
            facturas_por_pago = {}
            for d in detalles_pagos:
                p_id = d.get("id_pago")
                f_no = str(d.get("factura_no", "")).strip()
                if p_id and f_no:
                    facturas_por_pago.setdefault(p_id, []).append(f_no)

            # 2. Inicializar libro y estilos profesionales
            wb = openpyxl.Workbook()
            wb.remove(wb.active) # Eliminar hoja por defecto

            fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            font_title = Font(name="Calibri", size=14, bold=True, color="0F172A")
            font_sub = Font(name="Calibri", size=10, italic=True, color="64748B")
            font_total = Font(name="Calibri", size=11, bold=True, color="0F172A")
            fill_total = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

            num_fmt_curr = '"$"#,##0.00'
            num_fmt_qty = '#,##0.00'
            fecha_emision = datetime.datetime.now().strftime("%d/%m/%Y %I:%M %p")
            nombre_empresa = "DESECHABLES Y ABARROTES DOÑA MARY S.A.S."

            def agregar_encabezado(ws, titulo, headers, col_widths):
                ws['A1'] = nombre_empresa
                ws['A1'].font = font_title
                ws['A2'] = f"CONSOLIDADO MAESTRO: {titulo.upper()}"
                ws['A2'].font = Font(name="Calibri", size=12, bold=True, color="2563EB")
                ws['A3'] = f"Fecha de emisión: {fecha_emision} | Datos consolidados de todos los períodos"
                ws['A3'].font = font_sub
                ws.append([])
                ws.append(headers)
                for col_idx, width in enumerate(col_widths, start=1):
                    c_letter = get_column_letter(col_idx)
                    ws.column_dimensions[c_letter].width = width
                    cell = ws.cell(row=5, column=col_idx)
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # ----------------------------------------------------
            # HOJA 1: COMPRAS
            # ----------------------------------------------------
            ws_c = wb.create_sheet(title="Compras")
            c_headers = ["Fecha", "Código Insumo", "Nombre Insumo", "Proveedor", "Factura / Entrada", "Bodega", "Cantidad", "Costo Unitario", "IVA", "Costo Total"]
            c_widths = [14, 16, 35, 25, 20, 15, 14, 16, 14, 18]
            agregar_encabezado(ws_c, "Registro Completo de Compras y Entradas", c_headers, c_widths)

            for r_idx, c in enumerate(raw_compras, start=6):
                cat_i = c.get("catalogo_insumos") or {}
                ws_c.cell(row=r_idx, column=1, value=str(c.get("fecha", ""))[:10])
                ws_c.cell(row=r_idx, column=2, value=str(c.get("codigo_insumo", "")))
                ws_c.cell(row=r_idx, column=3, value=cat_i.get("nombre") or str(c.get("descripcion", "Insumo")))
                ws_c.cell(row=r_idx, column=4, value=str(c.get("proveedor", "Varios")))
                ws_c.cell(row=r_idx, column=5, value=str(c.get("numero_factura") or c.get("numero_entrada") or "S/N"))
                ws_c.cell(row=r_idx, column=6, value=str(c.get("bodega") or "Principal"))
                
                c7 = ws_c.cell(row=r_idx, column=7, value=float(c.get("cantidad") or 0))
                c7.number_format = num_fmt_qty
                c8 = ws_c.cell(row=r_idx, column=8, value=float(c.get("costo_unitario") or 0))
                c8.number_format = num_fmt_curr
                c9 = ws_c.cell(row=r_idx, column=9, value=float(c.get("valor_iva") or c.get("iva") or 0))
                c9.number_format = num_fmt_curr
                c10 = ws_c.cell(row=r_idx, column=10, value=float(c.get("costo_total") or 0))
                c10.number_format = num_fmt_curr

            last_c = len(raw_compras) + 5
            if raw_compras:
                tot_c = last_c + 1
                ws_c.cell(row=tot_c, column=1, value="TOTAL COMPRAS:").font = font_total
                c7_tot = ws_c.cell(row=tot_c, column=7, value=f'=SUM(G6:G{last_c})')
                c7_tot.font = font_total; c7_tot.number_format = num_fmt_qty
                c9_tot = ws_c.cell(row=tot_c, column=9, value=f'=SUM(I6:I{last_c})')
                c9_tot.font = font_total; c9_tot.number_format = num_fmt_curr
                c10_tot = ws_c.cell(row=tot_c, column=10, value=f'=SUM(J6:J{last_c})')
                c10_tot.font = font_total; c10_tot.number_format = num_fmt_curr
                for col_i in range(1, 11):
                    ws_c.cell(row=tot_c, column=col_i).fill = fill_total

            # ----------------------------------------------------
            # HOJA 2: VENTAS
            # ----------------------------------------------------
            ws_v = wb.create_sheet(title="Ventas")
            v_headers = ["Fecha", "Código Insumo", "Nombre Insumo", "Cliente", "Tipo Documento", "Factura / Remisión", "Cantidad", "Subtotal", "IVA", "Total Facturado"]
            v_widths = [14, 16, 35, 28, 18, 20, 14, 16, 14, 18]
            agregar_encabezado(ws_v, "Registro Completo de Ventas y Salidas", v_headers, v_widths)

            for r_idx, v in enumerate(raw_ventas, start=6):
                cat_i = v.get("catalogo_insumos") or {}
                ws_v.cell(row=r_idx, column=1, value=str(v.get("fecha", ""))[:10])
                ws_v.cell(row=r_idx, column=2, value=str(v.get("codigo_insumo", "")))
                ws_v.cell(row=r_idx, column=3, value=cat_i.get("nombre") or str(v.get("descripcion", "Insumo")))
                ws_v.cell(row=r_idx, column=4, value=str(v.get("cliente", "Clientes Varios")))
                ws_v.cell(row=r_idx, column=5, value=str(v.get("tipo_documento", "Factura POS")))
                ws_v.cell(row=r_idx, column=6, value=str(v.get("factura_no", "S/N")))
                
                v7 = ws_v.cell(row=r_idx, column=7, value=float(v.get("cantidad") or 0))
                v7.number_format = num_fmt_qty
                v8 = ws_v.cell(row=r_idx, column=8, value=float(v.get("subtotal") or 0))
                v8.number_format = num_fmt_curr
                v9 = ws_v.cell(row=r_idx, column=9, value=float(v.get("iva") or 0))
                v9.number_format = num_fmt_curr
                v10 = ws_v.cell(row=r_idx, column=10, value=float(v.get("total") or 0))
                v10.number_format = num_fmt_curr

            last_v = len(raw_ventas) + 5
            if raw_ventas:
                tot_v = last_v + 1
                ws_v.cell(row=tot_v, column=1, value="TOTAL VENTAS:").font = font_total
                v7_tot = ws_v.cell(row=tot_v, column=7, value=f'=SUM(G6:G{last_v})')
                v7_tot.font = font_total; v7_tot.number_format = num_fmt_qty
                v8_tot = ws_v.cell(row=tot_v, column=8, value=f'=SUM(H6:H{last_v})')
                v8_tot.font = font_total; v8_tot.number_format = num_fmt_curr
                v9_tot = ws_v.cell(row=tot_v, column=9, value=f'=SUM(I6:I{last_v})')
                v9_tot.font = font_total; v9_tot.number_format = num_fmt_curr
                v10_tot = ws_v.cell(row=tot_v, column=10, value=f'=SUM(J6:J{last_v})')
                v10_tot.font = font_total; v10_tot.number_format = num_fmt_curr
                for col_i in range(1, 11):
                    ws_v.cell(row=tot_v, column=col_i).fill = fill_total

            # ----------------------------------------------------
            # HOJA 3: AJUSTES
            # ----------------------------------------------------
            ws_a = wb.create_sheet(title="Ajustes")
            a_headers = ["Fecha", "Código Insumo", "Nombre Insumo", "Tipo", "Cantidad", "Costo Unitario", "Costo Total", "Motivo"]
            a_widths = [14, 16, 35, 14, 14, 16, 18, 30]
            agregar_encabezado(ws_a, "Registro de Ajustes de Inventario y Mermas", a_headers, a_widths)

            for r_idx, a in enumerate(raw_ajustes, start=6):
                cat_i = a.get("catalogo_insumos") or {}
                es_ent = str(a.get("tipo_ajuste", "")).upper() in ('AJUSTE_ENTRADA', 'ENTRADA_POR_SOBRANTE')
                ws_a.cell(row=r_idx, column=1, value=str(a.get("fecha_ajuste", ""))[:10])
                ws_a.cell(row=r_idx, column=2, value=str(a.get("codigo_insumo", "")))
                ws_a.cell(row=r_idx, column=3, value=cat_i.get("nombre", "Insumo"))
                ws_a.cell(row=r_idx, column=4, value="Entrada" if es_ent else "Salida")
                
                a5 = ws_a.cell(row=r_idx, column=5, value=float(a.get("cantidad") or 0))
                a5.number_format = num_fmt_qty
                a6 = ws_a.cell(row=r_idx, column=6, value=float(a.get("costo_unitario_congelado") or 0))
                a6.number_format = num_fmt_curr
                a7 = ws_a.cell(row=r_idx, column=7, value=float(a.get("costo_total_ajuste") or 0))
                a7.number_format = num_fmt_curr
                ws_a.cell(row=r_idx, column=8, value=str(a.get("motivo_observacion", "")))

            last_a = len(raw_ajustes) + 5
            if raw_ajustes:
                tot_a = last_a + 1
                ws_a.cell(row=tot_a, column=1, value="TOTAL AJUSTES:").font = font_total
                a5_tot = ws_a.cell(row=tot_a, column=5, value=f'=SUM(E6:E{last_a})')
                a5_tot.font = font_total; a5_tot.number_format = num_fmt_qty
                a7_tot = ws_a.cell(row=tot_a, column=7, value=f'=SUM(G6:G{last_a})')
                a7_tot.font = font_total; a7_tot.number_format = num_fmt_curr
                for col_i in range(1, 9):
                    ws_a.cell(row=tot_a, column=col_i).fill = fill_total

            # ----------------------------------------------------
            # HOJA 4: INVENTARIO (HOJA MAESTRA CON FÓRMULAS VINCULADAS)
            # ----------------------------------------------------
            ws_inv = wb.create_sheet(title="Inventario")
            inv_headers = [
                "Código", "Nombre", "Categoría", "Ubicación", "Stock Inicial",
                "Entradas Compras", "Costo Entradas", "Salidas Ventas", "Ingresos Salidas",
                "Stock Actual", "Costo Stock Actual", "Proyección Ingresos",
                "Costo Unitario", "Precio Venta",
                "Ajustes Entradas", "Ajustes Salidas", "Costo Ajustes Entradas", "Ingresos Ajustes Salidas"
            ]
            inv_widths = [16, 35, 18, 14, 14, 16, 18, 16, 18, 14, 18, 20, 16, 16, 16, 16, 18, 18]
            agregar_encabezado(ws_inv, "Catálogo Maestro de Inventario y Valorización Formulado", inv_headers, inv_widths)

            for idx, i in enumerate(raw_inv, start=6):
                code = str(i.get("codigo_insumo", ""))
                ws_inv.cell(row=idx, column=1, value=code)
                ws_inv.cell(row=idx, column=2, value=str(i.get("nombre", "")))
                ws_inv.cell(row=idx, column=3, value=str(i.get("categoria", "")))
                ws_inv.cell(row=idx, column=4, value=str(i.get("ubicacion") or "N/A"))
                
                e_val = ws_inv.cell(row=idx, column=5, value=float(i.get("stock_inicial") or 0))
                e_val.number_format = num_fmt_qty
                
                f_val = ws_inv.cell(row=idx, column=6, value=f'=SUMIF(Compras!B:B, A{idx}, Compras!G:G)')
                f_val.number_format = num_fmt_qty
                g_val = ws_inv.cell(row=idx, column=7, value=f'=SUMIF(Compras!B:B, A{idx}, Compras!J:J)')
                g_val.number_format = num_fmt_curr
                
                h_val = ws_inv.cell(row=idx, column=8, value=f'=SUMIF(Ventas!B:B, A{idx}, Ventas!G:G)')
                h_val.number_format = num_fmt_qty
                i_val = ws_inv.cell(row=idx, column=9, value=f'=SUMIF(Ventas!B:B, A{idx}, Ventas!J:J)')
                i_val.number_format = num_fmt_curr
                
                j_val = ws_inv.cell(row=idx, column=10, value=f'=E{idx}+F{idx}-H{idx}+O{idx}-P{idx}')
                j_val.number_format = num_fmt_qty
                k_val = ws_inv.cell(row=idx, column=11, value=f'=J{idx}*M{idx}')
                k_val.number_format = num_fmt_curr
                l_val = ws_inv.cell(row=idx, column=12, value=f'=J{idx}*N{idx}')
                l_val.number_format = num_fmt_curr
                
                m_val = ws_inv.cell(row=idx, column=13, value=float(i.get("costo_unitario") or 0))
                m_val.number_format = num_fmt_curr
                n_val = ws_inv.cell(row=idx, column=14, value=float(i.get("precio_venta") or 0))
                n_val.number_format = num_fmt_curr
                
                o_val = ws_inv.cell(row=idx, column=15, value=f'=SUMIFS(Ajustes!E:E, Ajustes!B:B, A{idx}, Ajustes!D:D, "Entrada")')
                o_val.number_format = num_fmt_qty
                p_val = ws_inv.cell(row=idx, column=16, value=f'=SUMIFS(Ajustes!E:E, Ajustes!B:B, A{idx}, Ajustes!D:D, "Salida")')
                p_val.number_format = num_fmt_qty
                q_val = ws_inv.cell(row=idx, column=17, value=f'=O{idx}*M{idx}')
                q_val.number_format = num_fmt_curr
                r_val = ws_inv.cell(row=idx, column=18, value=f'=P{idx}*N{idx}')
                r_val.number_format = num_fmt_curr

            last_inv = len(raw_inv) + 5
            if raw_inv:
                tot_inv = last_inv + 1
                ws_inv.cell(row=tot_inv, column=1, value="TOTAL GENERAL INVENTARIO:").font = font_total
                for c_idx, c_let, is_curr in [(5, 'E', False), (6, 'F', False), (7, 'G', True), (8, 'H', False), (9, 'I', True), (10, 'J', False), (11, 'K', True), (12, 'L', True), (15, 'O', False), (16, 'P', False), (17, 'Q', True), (18, 'R', True)]:
                    tot_cell = ws_inv.cell(row=tot_inv, column=c_idx, value=f'=SUM({c_let}6:{c_let}{last_inv})')
                    tot_cell.font = font_total
                    tot_cell.number_format = num_fmt_curr if is_curr else num_fmt_qty
                for col_i in range(1, 19):
                    ws_inv.cell(row=tot_inv, column=col_i).fill = fill_total

            # ----------------------------------------------------
            # HOJA 5: CARTERA (DETALLE DE DOCUMENTOS Y CRÉDITOS)
            # ----------------------------------------------------
            ws_cart = wb.create_sheet(title="Cartera")
            cart_headers = ["Cliente", "Encargado / Vendedor", "Tipo Documento", "Documento / Factura", "Fecha Emisión", "Total Factura", "Total Recaudado", "Saldo Faltante", "Estado"]
            cart_widths = [28, 20, 18, 20, 14, 18, 18, 18, 14]
            agregar_encabezado(ws_cart, "Libro de Cuentas por Cobrar y Documentos de Cartera", cart_headers, cart_widths)

            for idx, d in enumerate(docs_cartera, start=6):
                c_nom = str(d.get("cliente", ""))
                ws_cart.cell(row=idx, column=1, value=c_nom)
                ws_cart.cell(row=idx, column=2, value=vendedor_map.get(c_nom, "Sin Asignar"))
                ws_cart.cell(row=idx, column=3, value=str(d.get("tipo_documento", "Factura POS")))
                ws_cart.cell(row=idx, column=4, value=str(d.get("factura_no", "")))
                ws_cart.cell(row=idx, column=5, value=str(d.get("fecha", ""))[:10])
                
                c6 = ws_cart.cell(row=idx, column=6, value=float(d.get("total_factura") or 0.0))
                c6.number_format = num_fmt_curr
                c7 = ws_cart.cell(row=idx, column=7, value=float(d.get("total_abonado") or 0.0))
                c7.number_format = num_fmt_curr
                c8 = ws_cart.cell(row=idx, column=8, value=f'=F{idx}-G{idx}')
                c8.number_format = num_fmt_curr
                ws_cart.cell(row=idx, column=9, value=f'=IF(H{idx}<=0.01, "PAGADA", IF(G{idx}>0, "PARCIAL", "PENDIENTE"))')

            last_cart = len(docs_cartera) + 5
            if docs_cartera:
                tot_cart = last_cart + 1
                ws_cart.cell(row=tot_cart, column=1, value="TOTAL GENERAL CARTERA:").font = font_total
                for c_idx, c_let in [(6, 'F'), (7, 'G'), (8, 'H')]:
                    tot_cell = ws_cart.cell(row=tot_cart, column=c_idx, value=f'=SUM({c_let}6:{c_let}{last_cart})')
                    tot_cell.font = font_total
                    tot_cell.number_format = num_fmt_curr
                for col_i in range(1, 10):
                    ws_cart.cell(row=tot_cart, column=col_i).fill = fill_total

            # ----------------------------------------------------
            # HOJA 6: HISTORIAL PAGOS CARTERA
            # ----------------------------------------------------
            ws_pag = wb.create_sheet(title="Historial Pagos Cartera")
            pag_headers = ["Fecha Pago", "Cliente", "Encargado / Vendedor", "Método de Pago", "Banco / Entidad", "Referencia / Comprobante", "Monto Recaudado", "Facturas Aplicadas", "Observaciones", "Usuario Registro"]
            pag_widths = [18, 28, 20, 16, 18, 22, 18, 25, 25, 14]
            agregar_encabezado(ws_pag, "Historial Completo de Recaudos y Abonos de Cartera", pag_headers, pag_widths)

            for idx, p in enumerate(pagos_cartera, start=6):
                c_nom = str(p.get("nombre_cliente", ""))
                p_id = p.get("id_pago")
                facs_str = ", ".join([f"#{f}" for f in facturas_por_pago.get(p_id, [])]) or "Global (FIFO)"
                
                ws_pag.cell(row=idx, column=1, value=str(p.get("fecha_pago", ""))[:19].replace("T", " "))
                ws_pag.cell(row=idx, column=2, value=c_nom)
                ws_pag.cell(row=idx, column=3, value=vendedor_map.get(c_nom, "Sin Asignar"))
                ws_pag.cell(row=idx, column=4, value=str(p.get("metodo_pago", "EFECTIVO")))
                ws_pag.cell(row=idx, column=5, value=str(p.get("banco_origen") or "-"))
                ws_pag.cell(row=idx, column=6, value=str(p.get("referencia_comprobante") or "-"))
                
                p7 = ws_pag.cell(row=idx, column=7, value=float(p.get("monto_total") or 0.0))
                p7.number_format = num_fmt_curr
                ws_pag.cell(row=idx, column=8, value=facs_str)
                ws_pag.cell(row=idx, column=9, value=str(p.get("observaciones") or "-"))
                ws_pag.cell(row=idx, column=10, value=str(p.get("usuario_registro") or "admin"))

            last_pag = len(pagos_cartera) + 5
            if pagos_cartera:
                tot_pag = last_pag + 1
                ws_pag.cell(row=tot_pag, column=1, value="TOTAL RECAUDADO:").font = font_total
                tot_cell = ws_pag.cell(row=tot_pag, column=7, value=f'=SUM(G6:G{last_pag})')
                tot_cell.font = font_total
                tot_cell.number_format = num_fmt_curr
                for col_i in range(1, 11):
                    ws_pag.cell(row=tot_pag, column=col_i).fill = fill_total

            wb.save(file_path)

            if self.page:
                self.page.snack_bar = ft.SnackBar(ft.Text("¡Consolidado Excel generado con éxito en pocos segundos!"), bgcolor="green")
                self.page.snack_bar.open = True
                self.page.update()

        except Exception as ex:
            print(f"Error generando Excel: {ex}")
            if self.page:
                self.page.snack_bar = ft.SnackBar(ft.Text(f"Error al generar Excel: {ex}"), bgcolor="red")
                self.page.snack_bar.open = True
                self.page.update()
